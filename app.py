from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os, uuid, re
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)

# ── Config ────────────────────────────────────────────────────────────────────
BASE_UPLOAD_DIR = r'D:\VendorAuthSystem\uploads'
EXCEL_FILE      = r'D:\VendorAuthSystem\data\vendor_data.xlsx'
ALLOWED_EXT     = {'pdf','png','jpg','jpeg','doc','docx','xlsx','csv'}
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024

os.makedirs(BASE_UPLOAD_DIR, exist_ok=True)
os.makedirs(r'D:\VendorAuthSystem\data', exist_ok=True)

QUESTIONS = [
    {"id":1,"question":"Does your company have a valid GST registration certificate?",        "required_doc":"GST Certificate"},
    {"id":2,"question":"Does your company have ISO or any quality certification?",            "required_doc":"Quality Certificate"},
    {"id":3,"question":"Do you have a valid business license or trade license?",              "required_doc":"Business License"},
    {"id":4,"question":"Does your company have audited financial statements for last 2 yrs?", "required_doc":"Financial Statements"},
    {"id":5,"question":"Does your company have a PAN card or tax identification document?",   "required_doc":"PAN / Tax ID"},
    {"id":6,"question":"Do you have existing contracts or work orders from prior clients?",   "required_doc":"Work Order / Contract"},
]

sessions = {}

# ── Excel helpers ─────────────────────────────────────────────────────────────
def _border():
    s = Side(style='thin', color='2A3040')
    return Border(left=s, right=s, top=s, bottom=s)

def _init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Authentication"

    static = ["S.No","Vendor Name","Email","Session ID","Submitted At",
              "Yes Count","No Count","Files Uploaded"]
    q_cols = []
    for q in QUESTIONS:
        q_cols += [f"Q{q['id']} Answer", f"Q{q['id']} Document ({q['required_doc']})"]

    headers = static + q_cols
    hf = Font(name='Arial', bold=True, color="00E5B0", size=10)
    hb = PatternFill("solid", fgColor="1A1F2E")
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hb; c.alignment = ha; c.border = _border()

    widths = [6,28,30,36,22,12,12,16] + [14,34]*len(QUESTIONS)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'
    wb.save(EXCEL_FILE)

def _append_row(session):
    _init_excel()
    wb  = load_workbook(EXCEL_FILE)
    ws  = wb.active
    row = ws.max_row + 1
    sno = row - 1

    answers = session['answers']
    uploads = session['uploads']
    yes_c   = sum(1 for a in answers.values() if a['answer']=='yes')
    no_c    = sum(1 for a in answers.values() if a['answer']=='no')

    alt  = PatternFill("solid", fgColor="13171F" if sno%2==0 else "0F1218")
    bf   = Font(name='Arial', color="E8ECF0", size=10)
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    static_vals = [sno, session['vendor_name'], session['vendor_email'],
                   session['session_id'],
                   datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                   yes_c, no_c, len(uploads)]

    for col, val in enumerate(static_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = alt; c.font = bf; c.border = _border()
        c.alignment = lft if col in (2,3) else ctr

    for qi, q in enumerate(QUESTIONS):
        qid     = str(q['id'])
        ac_col  = 9  + qi*2
        fc_col  = 10 + qi*2
        answer  = answers.get(qid, {}).get('answer','N/A')
        relpath = uploads.get(qid, {}).get('relative_path','-')

        ac = ws.cell(row=row, column=ac_col, value=answer.upper())
        ac.border = _border(); ac.alignment = ctr
        if answer=='yes':
            ac.fill = PatternFill("solid", fgColor="0D2B22")
            ac.font = Font(name='Arial', bold=True, color="00CC88", size=10)
        elif answer=='no':
            ac.fill = PatternFill("solid", fgColor="2B1A0D")
            ac.font = Font(name='Arial', bold=True, color="FF6B35", size=10)
        else:
            ac.fill = alt; ac.font = bf

        fc = ws.cell(row=row, column=fc_col, value=relpath)
        fc.fill = alt; fc.border = _border(); fc.alignment = lft
        fc.font = Font(name='Arial', color="8A95A0", size=9, italic=(relpath=='-'))

    ws.row_dimensions[row].height = 22
    wb.save(EXCEL_FILE)

# ── Folder helper ─────────────────────────────────────────────────────────────
def _vendor_folder(vendor_name, session_id):
    safe = re.sub(r'[^\w\- ]', '', vendor_name).strip().replace(' ','_')
    name = f"{safe}_{session_id[:8]}"
    path = os.path.join(BASE_UPLOAD_DIR, name)
    os.makedirs(path, exist_ok=True)
    return path, name

def _allowed(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXT

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('../frontend', 'index.html')

@app.route('/api/questions')
def get_questions():
    return jsonify({"success":True,"total":len(QUESTIONS),"questions":QUESTIONS})

@app.route('/api/session/start', methods=['POST'])
def start_session():
    data  = request.get_json() or {}
    name  = data.get('vendor_name','').strip()
    email = data.get('vendor_email','').strip()
    if not name or not email:
        return jsonify({"success":False,"error":"Name and email required."}), 400

    sid = str(uuid.uuid4())
    folder_path, folder_name = _vendor_folder(name, sid)
    sessions[sid] = {
        "session_id":  sid, "vendor_name": name, "vendor_email": email,
        "folder_path": folder_path, "folder_name": folder_name,
        "answers": {}, "uploads": {}, "started_at": datetime.now().isoformat(),
        "completed": False,
    }
    return jsonify({"success":True,"session_id":sid,"message":f"Session started for {name}"})

@app.route('/api/upload', methods=['POST'])
def upload_file():
    sid = request.form.get('session_id')
    qid = request.form.get('question_id')
    if not sid or sid not in sessions:
        return jsonify({"success":False,"error":"Invalid session."}), 400
    if 'file' not in request.files:
        return jsonify({"success":False,"error":"No file provided."}), 400

    file = request.files['file']
    if not file.filename or not _allowed(file.filename):
        return jsonify({"success":False,"error":f"Allowed: {', '.join(ALLOWED_EXT)}"}), 400

    session      = sessions[sid]
    safe_name    = secure_filename(file.filename)
    dest_name    = f"Q{qid}_{safe_name}"
    dest_path    = os.path.join(session['folder_path'], dest_name)
    file.save(dest_path)

    size     = os.path.getsize(dest_path)
    rel_path = os.path.join('vendor_files', session['folder_name'], dest_name)

    session['uploads'][str(qid)] = {
        "original_name": safe_name, "saved_name": dest_name,
        "relative_path": rel_path,  "size_bytes":  size,
        "uploaded_at":   datetime.now().isoformat(),
    }
    return jsonify({"success":True,"message":"File uploaded.",
                    "file_name":safe_name,"file_size_kb":round(size/1024,2),
                    "saved_to":rel_path})

@app.route('/api/answer', methods=['POST'])
def save_answer():
    data   = request.get_json() or {}
    sid    = data.get('session_id')
    qid    = str(data.get('question_id'))
    answer = data.get('answer','').lower()
    if not sid or sid not in sessions:
        return jsonify({"success":False,"error":"Invalid session."}), 400
    if answer not in ('yes','no'):
        return jsonify({"success":False,"error":"Answer must be yes or no."}), 400
    if answer=='yes' and qid not in sessions[sid]['uploads']:
        return jsonify({"success":False,"error":"Upload document before answering Yes."}), 400
    sessions[sid]['answers'][qid] = {"answer":answer,"answered_at":datetime.now().isoformat()}
    return jsonify({"success":True,"question_id":qid,"answer":answer})

@app.route('/api/submit', methods=['POST'])
def submit_form():
    data = request.get_json() or {}
    sid  = data.get('session_id')
    if not sid or sid not in sessions:
        return jsonify({"success":False,"error":"Invalid session."}), 400
    session = sessions[sid]
    if session['completed']:
        return jsonify({"success":False,"error":"Already submitted."}), 400

    answered = set(session['answers'].keys())
    missing  = set(str(q['id']) for q in QUESTIONS) - answered
    if missing:
        return jsonify({"success":False,
                        "error":f"Missing answers: Q{sorted(int(m) for m in missing)}"}), 400

    for qid, ans in session['answers'].items():
        if ans['answer']=='yes' and qid not in session['uploads']:
            q   = next((x for x in QUESTIONS if str(x['id'])==qid),None)
            doc = q['required_doc'] if q else f"Q{qid}"
            return jsonify({"success":False,"error":f"Document missing for '{doc}'."}), 400

    session['completed']    = True
    session['completed_at'] = datetime.now().isoformat()
    _append_row(session)   # ← write to master Excel

    yes_c = sum(1 for a in session['answers'].values() if a['answer']=='yes')
    no_c  = sum(1 for a in session['answers'].values() if a['answer']=='no')

    return jsonify({
        "success": True,
        "message": "Submitted successfully! Row added to vendor_data.xlsx",
        "excel_file":    os.path.abspath(EXCEL_FILE),
        "vendor_folder": os.path.abspath(session['folder_path']),
        "summary": {
            "vendor_name":        session['vendor_name'],
            "total_questions":    len(QUESTIONS),
            "yes_answers":        yes_c,
            "no_answers":         no_c,
            "documents_uploaded": len(session['uploads']),
        }
    })

@app.route('/api/session/<sid>')
def get_session(sid):
    if sid not in sessions:
        return jsonify({"success":False,"error":"Not found."}), 404
    s = sessions[sid]
    return jsonify({"success":True,"session":{
        "vendor_name":s['vendor_name'],"vendor_email":s['vendor_email'],
        "folder":s['folder_name'],"answers":s['answers'],
        "uploads":{k:v['relative_path'] for k,v in s['uploads'].items()},
        "completed":s['completed'],
    }})

if __name__ == '__main__':
    _init_excel()
    print("🚀  Server          →  http://localhost:5000")
    print(f"📊  Excel file      →  D:\\VendorAuthSystem\\data\\vendor_data.xlsx")
    print(f"📁  Vendor uploads  →  D:\\VendorAuthSystem\\uploads\\")
    app.run(debug=True, host='0.0.0.0', port=5000)